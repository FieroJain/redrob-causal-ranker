#!/usr/bin/env python3
"""
Convert submissionv12.csv to the XLSX format required by the portal.
Run: python convert_to_xlsx.py
Requires: pip install openpyxl
"""

import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

INPUT_CSV = "submissionv12.csv"
OUTPUT_XLSX = "submission_final.xlsx"

def convert():
    wb = Workbook()
    ws = wb.active
    ws.title = "Ranked Candidates"

    with open(INPUT_CSV, encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = list(reader)

    header = rows[0]
    data_rows = rows[1:]

    # Write header with formatting
    for col_idx, col_name in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data rows
    for row_idx, row in enumerate(data_rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            # rank and score as numbers, not text
            if col_idx == 2:  # rank
                cell = ws.cell(row=row_idx, column=col_idx, value=int(value))
            elif col_idx == 3:  # score
                cell = ws.cell(row=row_idx, column=col_idx, value=float(value))
                cell.number_format = "0.0000"
            else:
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx == 4))

    # Column widths
    widths = {1: 18, 2: 8, 3: 10, 4: 100}
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT_XLSX)
    print(f"Written: {OUTPUT_XLSX}")
    print(f"Rows: {len(data_rows)} candidates + 1 header")

if __name__ == "__main__":
    convert()
